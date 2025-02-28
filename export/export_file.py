import calendar
import os
import numpy as np
import openpyxl as op
from win32com.client import Dispatch

def create_and_change_path(endperiod,folder_path=None, currency=None, money=None):
    file_path="Output"+"/"+"R1_"+str(endperiod)+"/"+folder_path+"/"+currency+"/"+money
    if not os.path.exists(file_path):
        os.makedirs(file_path)
    os.chdir(file_path)

def denotesymbol(data, datatype):
    if datatype=='fig':
        data[(data<0.5) & (data>0)] ='*'
        data.replace([np.nan, 0], '-',inplace=True)

    elif datatype=='share':
        data[(data<0.05) & (data>0)] ='*'
        data.replace([np.inf, -np.inf], '∞',inplace=True)
        data.replace([np.nan, 0], '-',inplace=True)

    elif datatype=='chg':
        data[((data<0.05) & (data>0))|((data>-0.05) & (data<0))] = 0.001
        data[((data>1000)&(data<np.inf))|((data<-1000)&(data>-np.inf))] = '..'
        data.replace([np.inf, -np.inf], '∞',inplace=True)
        data.replace([np.nan, 0], '-',inplace=True)
        data.replace(0.001, '*',inplace=True)
    return data

def money_conversion(data,currency,money):
    dollar = {'HKD':1, 'USD':7.8}
    unit = {'TH':1000, 'MN':1000000}
    result = data/dollar[currency]/unit[money]
    return result

def part1_toexcel_generaltrade(fig, chg, writer, currency, money):
    # convert money by currency in only figures
    fig_c = money_conversion(fig, currency, money)
    # denote symbols to the figures
    fig_d = denotesymbol(fig_c, datatype='fig')
    # export figures and change to excel
    fig_d.to_excel(writer,sheet_name=f"{currency}_{money}",index=False,startrow=5, startcol=3,header=False)
    chg.to_excel(writer,sheet_name=f"{currency}_{money}",index=False,startrow=5, startcol=7,header=False)

def part2a_toexcel_specialtrade_fig(fig, writer, currency, money, startrow):
    # convert money by currency in only figures
    fig_c = money_conversion(fig, currency, money)
    # denote symbols to the figures
    fig_d = denotesymbol(fig_c, datatype='fig')
    # export trade figures in different rows and columns
    fig_d.iloc[:,0].to_excel(writer,sheet_name=f"{currency}_{money}",index=False,startrow=startrow, startcol=2,header=False)
    fig_d.iloc[:,1].to_excel(writer,sheet_name=f"{currency}_{money}",index=False,startrow=startrow, startcol=4,header=False)
    fig_d.iloc[:,2].to_excel(writer,sheet_name=f"{currency}_{money}",index=False,startrow=startrow, startcol=6,header=False)
    fig_d.iloc[:,3].to_excel(writer,sheet_name=f"{currency}_{money}",index=False,startrow=startrow, startcol=8,header=False)

def part2b_toexcel_specialtrade_share(share, writer, currency, money, startrow):
    # export trade shares in different rows and columns
    share.iloc[:,0].to_excel(writer,sheet_name=f"{currency}_{money}",index=False,startrow=startrow, startcol=3,header=False)
    share.iloc[:,1].to_excel(writer,sheet_name=f"{currency}_{money}",index=False,startrow=startrow, startcol=5,header=False)
    share.iloc[:,-2].to_excel(writer,sheet_name=f"{currency}_{money}",index=False,startrow=startrow, startcol=7,header=False)
    share.iloc[:,-1].to_excel(writer,sheet_name=f"{currency}_{money}",index=False,startrow=startrow, startcol=9,header=False)

def part2c_toexcel_specialtrade_chg(chg, writer, currency, money, startrow):
    # export trade changes in different rows and columns
    chg.iloc[:,-1].to_excel(writer,sheet_name=f"{currency}_{money}",index=False,startrow=startrow, startcol=10,header=False)

def part3_toexcel_ranking_cty(ranking, writer, currency, money, periods):
    ranking.to_excel(writer, sheet_name=f"{currency}_{money}",index=False,
                         startrow=5, startcol=1,header=False)
    workbook  = writer.book
    worksheet = writer.sheets[f"{currency}_{money}"]
    fmt_bold = workbook.add_format({'bold': 1,'align': 'center','font_name': 'Arial',
                                        'font_size':7.5})
    fmt_bold_right = workbook.add_format({'bold': 1,'align': 'right','font_name': 'Arial',
                                        'font_size':7.5})
    #acquire latest month name
    month_name = calendar.month_abbr[int(periods[-1][-2:])]

    worksheet.write("C3", 'RANKING', fmt_bold)
    worksheet.write("C4", "JAN-"+month_name.upper(), fmt_bold)
    worksheet.write("B5", "'"+periods[-2][2:4], fmt_bold_right)
    worksheet.write("C5", "'"+periods[-1][2:4], fmt_bold_right)

def adjust_excelformat_xlsxwriter(writer, currency, money, periods, name, noofprod):
    workbook  = writer.book
    worksheet = writer.sheets[f"{currency}_{money}"]
    title = f"HONG KONG'S TOP {noofprod} TRADE WITH "+ name
    # setting format, labels, title, annotation for cells in the excel file
    merge_format_T = workbook.add_format({'bold': 1,'align': 'center',
                    'font_name': 'Arial','font_size':10})

    fmt_right = workbook.add_format({'align': 'right','font_name': "Arial",
                            'font_size':7.5})

    fmt_left = workbook.add_format({'align': 'left','font_name': "Arial",
                                'font_size':7.5})

    fmt_bold = workbook.add_format({'bold': 1,'align': 'center','font_name': 'Arial',
                                    'font_size':7.5})

    fmt_bold_left = workbook.add_format({'bold': 1,'align': 'left','font_name': 'Arial',
                                        'font_size':7.5})

    wrap_text = workbook.add_format({'font_name': 'Arial','font_size':7.5,
                                    'text_wrap':True})

    worksheet.set_column('A:A', 22, None)
    worksheet.set_column('B:B', 45, wrap_text)
    worksheet.set_column('C:K', None, fmt_right)
    worksheet.write("A6", 'TOTAL EXPORTS', fmt_bold_left)
    worksheet.write("A7", 'DOMESTIC EXPORTS', fmt_bold_left)
    worksheet.write("A8", 'RE-EXPORTS', fmt_bold_left)
    worksheet.write("A9", 'IMPORTS', fmt_bold_left)
    worksheet.write("A10", '(OF WHICH RE-EXPORTED)', fmt_bold_left)
    worksheet.write("A11", 'TOTAL TRADE', fmt_bold_left)
    worksheet.write("A12", 'TRADE BALANCE', fmt_bold_left)

    worksheet.write(16, 1, '-TOTAL EXPORTS-', fmt_bold)
    worksheet.write(16+noofprod+3, 1, '-DOMESTIC EXPORTS-', fmt_bold)
    worksheet.write(16+2*(noofprod+3), 1, '-RE-EXPORTS-', fmt_bold)
    worksheet.write(16+3*(noofprod+3), 1, '-IMPORTS-', fmt_bold)

    worksheet.merge_range('A1:J1', title, merge_format_T)
    worksheet.merge_range("H2:J2", f'VALUE : {currency} {money}', fmt_bold)
    worksheet.merge_range("H3:J3", '% CHANGE', fmt_bold)
    worksheet.merge_range("A14:B14", 'MAJOR COMMODITIES OF TRADES', fmt_bold)
    worksheet.write("A16", "SITC", fmt_bold)

    # writing VALUE and % SHARE labels
    for c in range(2,10,2):
        worksheet.write(15, c, "VALUE", fmt_bold)
        worksheet.write(15, c+1, "% SHARE", fmt_bold)
    worksheet.write(15, 10, "% CHG", fmt_bold)

    # year
    unique_yr = sorted(set([yr[:4]for yr in periods]))
    for i, yr in enumerate(unique_yr):
        worksheet.write(4, 3+i, int(yr), fmt_bold)
    worksheet.merge_range("C15:D15", int(unique_yr[0]), fmt_bold)
    worksheet.merge_range("E15:F15", int(unique_yr[1]), fmt_bold)
    worksheet.merge_range("G15:H15", int(unique_yr[2]), fmt_bold)
    worksheet.merge_range("I15:K15", int(unique_yr[3]), fmt_bold)

    worksheet.write("H5", f"'{unique_yr[1][-2:]}/{unique_yr[0][-2:]}", fmt_bold)
    worksheet.write("I5", f"'{unique_yr[2][-2:]}/{unique_yr[1][-2:]}", fmt_bold)
    worksheet.write("J5", f"'{unique_yr[3][-2:]}/{unique_yr[2][-2:]}", fmt_bold)

    # acquire latest month name
    month_name = calendar.month_abbr[int(periods[-1][-2:])]

    worksheet.write("G4", "JAN-"+month_name.upper(), fmt_bold)
    worksheet.write("J4", "JAN-"+month_name.upper(), fmt_bold)
    worksheet.merge_range("I14:K14", "JAN-"+month_name.upper(), fmt_bold)

    # hide gridlines
    worksheet.hide_gridlines(2)

    # write source, symbol annotation at the end
    worksheet.merge_range(18+4*(noofprod+3),0,18+4*(noofprod+3),5, "* INSIGNIFICANT            ∞ INFINITY", fmt_left)
    worksheet.merge_range(19+4*(noofprod+3),0,19+4*(noofprod+3),5, "..OVER 1000% INCREASE      - NIL     N.E.S. NOT ELSEWHERE SPECIFIED", fmt_left)
    worksheet.merge_range(20+4*(noofprod+3),0,20+4*(noofprod+3),5, "SOURCE: HONG KONG TRADE STATISTICS, CENSUS & STATISTICS DEPT.", fmt_left)
    worksheet.merge_range(20+4*(noofprod+3),6,20+4*(noofprod+3),10, "HONG KONG TRADE DEVELOPMENT COUNCIL", fmt_left)

def adjust_excelformat_openpyxl(excel_name, currency, money):
    wb = op.load_workbook(excel_name)
    ws = wb[f"{currency}_{money}"]

    # use openpyxl
    ft1 = op.styles.Font(name='Arial', size=7.5)
    value_format = '#,##0'
    pct_format = '#,##0.0'

    # no border format
    side = op.styles.Side(border_style=None)
    no_border = op.styles.borders.Border(
        left=side,
        right=side,
        top=side,
        bottom=side,
    )
    # row and column number start from 1, not 0 index
    ##### set upper part value format
    # set D6:G12 value format
    for row in range(6,13):
        for col in range(4,8):
            _cell = ws.cell(row,col)
            _cell.number_format = value_format
            _cell.font = ft1

    # set H6:J11 %CHG format
    for row in range(6,13):
        for col in range(8,11):
            _cell = ws.cell(row,col)
            _cell.font = ft1

    # set A:A SITC code format, up to row no.1000
    for row in range(18,1000):
        _cell = ws.cell(row,1)
        _cell.border = no_border
        _cell.font = ft1
        _cell.alignment = op.styles.Alignment(horizontal='left')
        _cell.number_format = '@'

    # set C:I value format, up to row no.1000
    for row in range(18,1000):
        for col in range(3,10,2):
            _cell = ws.cell(row,col)
            _cell.number_format = value_format
            _cell.font = ft1

    # set D:K %CHG format, up to row no.1000
    for row in range(18,1000):
        for col in [4,6,8,10,11]:
            _cell = ws.cell(row,col)
            _cell.font = ft1
            _cell.number_format = pct_format

    # set upper trade fig %CHG format
    for row in range(6,12):
        for col in range(8,11):
            _cell = ws.cell(row,col)
            _cell.font = ft1
            _cell.number_format = pct_format

    wb.save(excel_name)

def autofit(excel_name, currency, money):
    """using win32com.client to control excel to autofit"""
    excel = Dispatch('Excel.Application')
    thisdir = os.getcwd()
    wb = excel.Workbooks.Open(thisdir+"/"+excel_name)
    ws = wb.Worksheets(f"{currency}_{money}")
    ws.Columns.AutoFit()
    wb.Save()
    wb.Close()

def autofit_wrap_industry(excel_name):
    """using win32com.client to control excel to autofit"""
    #print("here test")
    #print(excel_name)
    excel = Dispatch('Excel.Application')
    thisdir = os.getcwd()
    #print("dir: ", thisdir)
    wb = excel.Workbooks.Open(thisdir+"/"+excel_name)
    #print('wb: ',wb)
    ws = wb.Worksheets
    ws_len = len(wb.Worksheets)
    #print(ws_len)

    for s in range(ws_len):
        #print(s)
        # wrap
        ws[s].Rows[1].WrapText = True
        # auto fit
        ws[s].Columns.AutoFit()

    wb.Save()
    wb.Close()

def freezepane(excel_name):
    wb = op.load_workbook(excel_name)

    i=0
    for ws in wb:
        # freeze_pane
        if i == 0: ws.freeze_panes = ws['B3']
        else: ws.freeze_panes = ws['B4']
        i+=1
    wb.save(excel_name)

def addcomma_align(excel_name):
    wb = op.load_workbook(excel_name)

    for ws in wb:
        # add comma format
        for cell in ws['B:B']:
            cell.number_format = '#,###0.0'
        for cell in ws['C:C']:
            cell.number_format = '#,###0.0'
        for cell in ws['E:E']:
            cell.number_format = '#,###0.0'
        for cell in ws['F:F']:
            cell.number_format = '#,###0.0'
        for cell in ws['H:H']:
            cell.number_format = '#,###0.0'

        # last row
        endrow = len(ws['A:A'])

        # align to right
        for row in range(3, endrow+1):
            for col in range(2,15):
                _cell = ws.cell(row,col)
                _cell.alignment = op.styles.Alignment(horizontal='right')

    wb.save(excel_name)

def addtitle(excel_name, industryname, EU_name="E.U. (excl. UK)"):
    # find the hyphen position and delete it
    positionhyphen = industryname[:8].find('-')
    if positionhyphen!=-1: industryname=industryname[positionhyphen+2:]

    wb = op.load_workbook(excel_name)
    # font format
    ft1 = op.styles.Font(name='Calibri', size=12.5, bold=True)
    #_cell.font = ft1
    #_cell.alignment = op.styles.Alignment(horizontal='left')
    # titles
    titles = [f'Performance of Hong Kong’s Trade: {industryname}', #1
              f"Hong Kong's Domestic Exports of {industryname} by Country", #2
              f"Hong Kong's Total Exports of {industryname} by Country", #3
              f"Hong Kong's Re-exports of {industryname} by Country as Destination", #4
              f"Hong Kong's Re-exports of {industryname} by Country as Origin", #4
              f"Hong Kong's Imports of {industryname} by Country as Consignment", #5
              f"Hong Kong's Imports of {industryname} by Country as Origin", #6
              f"Hong Kong's Total Trades of {industryname} by Country", #3

              f"Hong Kong's Domestic Exports of {industryname} by Country by Quantity", #7
              f"Hong Kong's Total Exports of {industryname} by Country by Quantity", #8
              f"Hong Kong's Re-exports of {industryname} by Country as Destination by Quantity", #9
              f"Hong Kong's Re-exports of {industryname} by Country as Origin by Quantity", #9
              f"Hong Kong's Imports of {industryname} by Country as Consignment by Quantity", #10
              f"Hong Kong's Imports of {industryname} by Country as Origin by Quantity", #11
              f"Hong Kong's Total Trades of {industryname} by Country by Quantity", #8
              # EU
              f"Hong Kong's Domestic Exports of {industryname} by {EU_name}", #12
              f"Hong Kong's Total Exports of {industryname} by {EU_name}", #13
              f"Hong Kong's Re-exports of {industryname} by {EU_name} as Destination", #4
              f"Hong Kong's Re-exports of {industryname} by {EU_name} as Origin", #4
              f"Hong Kong's Imports of {industryname} by {EU_name} as Consignment", #5
              f"Hong Kong's Imports of {industryname} by {EU_name} as Origin", #6
              f"Hong Kong's Total Trades of {industryname} by {EU_name}", #3

              # Asean
              f"Hong Kong's Domestic Exports of {industryname} by Asean", #14
              f"Hong Kong's Total Exports of {industryname} by Asean", #15
              f"Hong Kong's Re-exports of {industryname} by Asean as Destination", #4
              f"Hong Kong's Re-exports of {industryname} by Asean as Origin", #4
              f"Hong Kong's Imports of {industryname} by Asean as Consignment", #5
              f"Hong Kong's Imports of {industryname} by Asean as Origin", #6
              f"Hong Kong's Total Trades of {industryname} by Asean", #3

              # Asia
              f"Hong Kong's Domestic Exports of {industryname} by Asia", #16
              f"Hong Kong's Total Exports of {industryname} by Asia", #17

              # Europe
              f"Hong Kong's Imports of {industryname} by Europe by Country as Origin", #18

              # product
              f"Hong Kong's Domestic Exports of {industryname} by products", #19
              f"Hong Kong's Total Exports of {industryname} by products", #20
              f"Hong Kong's Re-exports of {industryname} by products", #21
              f"Hong Kong's Imports of {industryname} by products", #22
            ]

    for i, (ws, t) in enumerate(zip(wb, titles)):
        # first sheet has fewer columns
        if i == 0: ws.merge_cells('A1:I1')
        else: ws.merge_cells('A1:N1')

        ws.cell(row=1, column=1).value = t
        ws.cell(row=1, column=1).alignment = op.styles.Alignment(horizontal='center')
        ws.cell(row=1, column=1).font = ft1
    wb.save(excel_name)

def addsource(excel_name):
    wb = op.load_workbook(excel_name)
    for ws in wb:
        endrow = len(ws['A:A'])
        #print(endrow)
        ws.cell(row=endrow+2, column=1).value = '* INSIGNIFICANT            ∞ INFINITY'
        ws.cell(row=endrow+3, column=1).value = '..OVER 1000% INCREASE      - NIL     N.E.S. NOT ELSEWHERE SPECIFIED'
        ws.cell(row=endrow+4, column=1).value = 'SOURCE: HONG KONG TRADE STATISTICS, CENSUS & STATISTICS DEPT.'

    wb.save(excel_name)
